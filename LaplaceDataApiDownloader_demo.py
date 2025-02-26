# =============================================================================
#    ラプラス発電・売電データApi取得プログラム
#    作成者： SY.LEE
#    2023/02/24 ver.1.0 SY.LEE 初版作成 
#    2023/03/01 ver.1.2 SY.LEE 再帰処理追加
#    2023/03/02 ver.1.3 SY.LEE 計測機器の分岐処理追加、ラプラス以外は処理対
#    2023/03/07 ver.1.5 SY.LEE PCSデータ取得機能追加
#    2023/03/10 ver.1.6 SY.LEE エナジーソリューションデータ取得機能追加
#    2023/04/10 ver.1.7 SY.LEE 日付入力の上限を変更
#    2023/07/04 ver.1.8 SY.LEE 複数案件及び30分単位以外も取得可能に改修
#    2023/08/24 ver.1.9 SY.LEE エナジーソリューションズ URL変更に伴うAPI URL修正
#    2024/03/26 ver.2.0 kakitani.r エナジーソリューションズのデータをseleniumを使って取得するよう修正
#    2024/08/26 ver.2.1 kakitani.r エナジーソリューションズのデータをseleniums仕様変更に伴うコード修正
#    2024/08/26 ver.2.2 kakitani.r ファイル名に採番を追加
#    2024/11/06 ver.2.3 kakitani.r ラプラス実行後の追加メニューにエナソリュ追加
#    2025/1/09 ver.2.4 kakitani.r エナソリュのDLボタンを日締め→監視データに変更

# =============================================================================

import pymysql
import requests
import pandas
from dotenv import load_dotenv
from requests.auth import HTTPDigestAuth
from datetime import datetime, timedelta
from selenium.webdriver.common.by import By
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.select import Select
import shutil
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, PatternFill, Color, Border, Side
from openpyxl.styles.fonts import Font
import os
import sys
import glob
import json
import time
import warnings
import logging



# define Local Variable
DOWNLOAD_PATH = 'C:\\python\\download\\'
IMP_PATH = 'C:\\python\\imp\\'
GET_MULTI_TARGET_PATH = 'C:\\python\\target\\'
ENERGYSOLUTIONS_DOWNDOAD_FILE = 'C:\\python\\download\\chrome_download\\'


SMARTMETER_TYPE = 'approvedmeter'
PCS_TYPE = 'pcs'

ECOMEGANE_CODE = 10
LAPLACE_CODE = 20
ENERGYSOLUTIONS_CODE = 30
MONTHLY_DAYS = 31

LAPLACE_CODE_NAME = "ラプラス"
ENERGYSOLUTIONS_CODE_NAME = "エナジーソリューションズ"

PCS_TYPE_JAP = "パワーコンディショナー"
SMARTMETER_TYPE_JAP = "スマートメーター"

DATA_UNIT_MINUTE_JAP = "1分"
DATA_UNIT_HALF_HOUR_JAP = "30分"
DATA_UNIT_HOUR_JAP = "1時間"
DATA_UNIT_DAY_JAP = "1日"
DATA_UNIT_MONTH_JAP = "1ヶ月"

# api parammeter
MINUTE = "minute"
HALF_HOUR = "halfhour"
HOUR = "hour"
DAY = "day"
MONTH = "month"

load_dotenv('LaplaceDataApiDownloader.env')
warnings.simplefilter(action='ignore', category=FutureWarning)

log_file = 'LaplaceDataApiDownloader.txt'
logger = logging.getLogger(__name__)  # 名前を指定してロガーを取得
logger.setLevel(logging.ERROR)
file_handler = logging.FileHandler(log_file, mode='a')
formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
file_handler.setFormatter(formatter)
logger.addHandler(file_handler)


def download() :
    if not os.path.exists(IMP_PATH) :
        os.makedirs(IMP_PATH)
    print("\n【ラプラス・エナジーソリューションズ発電データAPI取得プログラム】")
    print("\n複数案件を取得する場合は「y」を、１案件のみ取得する場合は「n」を入力。")

    getMultiTarget = input("\n ・「y」 or 「n」を入力 : ")

    if getMultiTarget == "y" or getMultiTarget == "Y" :
        print("\n【複数案件取得手順１】自身の「C:\\python\\target」にアクセスする。（プログラムで既に作成されている）")
        print("【複数案件取得手順２】取得する対象の物件番号を「【Template】物件番号リスト.xlsx」にまとめ、「C:\\python\\target」に入れる。")
        print("【複数案件取得手順３】物件番号リストファイルに暗号化がかかっていないことを確認。")

        getMultiFlg = True

    else :
        getMultiFlg = False

    # ただし、複数案件だった物件番号リストで処理するように変更
    excelData = []
    targetObjectNoList = []
    targetStartDateList = []
    targetEndDateList = []

    if getMultiFlg == True :

        # 複数案件読み取り処理
        df = pandas.read_excel(GET_MULTI_TARGET_PATH + "【Template】物件番号リスト.xlsx")

        for key in df.keys() :
            excelData.append(df[key])

        targetObjectNoList.append(excelData[0])
        targetStartDateList.append(excelData[1])
        targetEndDateList.append(excelData[2])

        # ループカウント設定
        maxIndex = len(targetObjectNoList[0])

    else :
        print("\n【Tip】物件番号はコピー＆マウス右クリックでツール入力欄に貼り付け可能。")
        targetObjectNoList.append(input("\n ・物件番号 : "))

        # ループカウント設定
        maxIndex = len(targetObjectNoList)

    print("\nダウンロードする計測機器を選択してください。")
    print("\n ・0 : ラプラス")
    print(" ・1 : エナジーソリューションズ")

    inputMeasuringEquip = input("\n ・計測機器 : ")

    if inputMeasuringEquip == '0' :
        inputMeasuringEquip = LAPLACE_CODE

        print("\nダウンロードデータタイプ")
        print("\n ・0 : スマートメーター")
        print(" ・1 : PCS")

        inputDataType = input("\n ・データタイプ : ")

        if inputDataType == '0' :
            dataType = SMARTMETER_TYPE

            print("\n データ単位を入力。")
            print("\n ・0 : 30分")
            print(" ・1 : 1時間")
            print(" ・2 : 1日")
            print(" ・3 : 1ヶ月")

            inputDataUnit = input("\n ・データ単位 : ")

            if inputDataUnit == "0" :
                unit = HALF_HOUR
            elif inputDataUnit == "1" :
                unit = HOUR
            elif inputDataUnit == "2" :
                unit = DAY
            elif inputDataUnit == "3" :
                unit = MONTH
            else :
                print("\nデータタイプ入力誤り。最初からやり直してください。")
                download()

        elif inputDataType == '1' :
            dataType = PCS_TYPE

            print("\n データ単位を入力。")
            print("\n ・0 : 1分")
            print(" ・1 : 30分")

            inputDataUnit = input("\n ・データ単位 : ")

            if inputDataUnit == "0" :
                unit = MINUTE
            elif inputDataUnit == "1" :
                unit = HALF_HOUR
            else :
                print("\nデータタイプ入力誤り。最初からやり直してください。")
                download()
        else :
            print("\nデータタイプ入力誤り。最初からやり直してください。")
            download()

    elif inputMeasuringEquip == '1' :
        inputMeasuringEquip = ENERGYSOLUTIONS_CODE
        # EnergySolutionでは使用しませんが、一度設定
        unit = DAY
        dataType = SMARTMETER_TYPE

    else :
        print("\n計測機器タイプ入力誤り。最初からやり直してください。")
        download()

    if getMultiFlg == False :

        if unit == MONTH :
            print("\n※ 月単位の取得期間入力例) : 202201")
        else :
            print("\n※ 取得期間入力例) : 20220101")

        targetStartDateList.append(input("\n ・開始期間 : "))
        targetEndDateList.append(input(" ・終了期間 : "))

    # データオブジェクト数だけAPIを実行する
    for index in range(maxIndex) :

        if getMultiFlg == True :
            targetObjectNo = str(targetObjectNoList[0][index])
            targetStartDate = str(targetStartDateList[0][index])
            targetEndDate = str(targetEndDateList[0][index])
        else :
            targetObjectNo = targetObjectNoList[0]
            targetStartDate = targetStartDateList[0]
            targetEndDate = targetEndDateList[0]

        try :
            if unit == MONTH :
                # string->date변환
                targetStartDateToDate = datetime.strptime(targetStartDate, "%Y%m")
                targetEndDateToDate = datetime.strptime(targetEndDate, "%Y%m")
            else :
                # string->date변환
                targetStartDateToDate = datetime.strptime(targetStartDate, "%Y%m%d")
                targetEndDateToDate = datetime.strptime(targetEndDate, "%Y%m%d")

        except ValueError :
            print("\n日付形式エラー。最初からやり直してください。")
            # 再帰処理
            download()

        # 開始日と終了日の妥当性チェック1
        if targetStartDateToDate > targetEndDateToDate :
            print("\n開始日は終了日より過去日を入力してください。")
            # 再帰処理
            download()

        # 開始日と終了日の妥当性チェック2
        if unit == MONTH and int(targetStartDate) - int(targetEndDate) <= -100 :
            print("\nデータ単位が月の場合、最大取得範囲は12ヶ月です。（ラプラス側の仕様）")
            # 再帰処理
            download()

        # Create download data info

        print("\n　◆ 物件番号 : " + targetObjectNo)

        if inputMeasuringEquip == LAPLACE_CODE :
            print("　◆ 計測機器 : " + LAPLACE_CODE_NAME)
        else :
            print("　◆ 計測機器 : " + ENERGYSOLUTIONS_CODE_NAME)

        if dataType == PCS_TYPE :
            print("　◆ データタイプ : " + PCS_TYPE_JAP)
        else :
            print("　◆ データタイプ : " + SMARTMETER_TYPE_JAP)

        if unit == MINUTE :
            print("　◆ データ単位 : " + DATA_UNIT_MINUTE_JAP)
        elif unit == HALF_HOUR :
            print("　◆ データ単位 : " + DATA_UNIT_HALF_HOUR_JAP)
        elif unit == HOUR :
            print("　◆ データ単位 : " + DATA_UNIT_HOUR_JAP)
        elif unit == DAY :
            print("　◆ データ単位 : " + DATA_UNIT_DAY_JAP)
        else :
            print("　◆ データ単位 : " + DATA_UNIT_MONTH_JAP)

        if unit == MONTH :
            print("　◆ 取得期間 : " + targetStartDate + " ～ " + targetEndDate)
        else :
            # Date差の取得
            date_diff = targetEndDateToDate - targetStartDateToDate
            print("　◆ 取得期間 : " + targetStartDate + " ～ " + targetEndDate + "（日数:" + str(date_diff.days +1) + "）")

        # create Database connection
        db = pymysql.connect(
            host = '',
            port = '',
            user = '',
            passwd = '',
            db = '',
            charset = '')

        try :
            with db.cursor() as cursor :
                # Select API Key1,2 from EP System database.
                sql = """
                    SELECT
                        pt_solar_measuring_info.pk_measuring_equip_id,
                        pt_solar_measuring_info.api_key1,
                        pt_solar_measuring_info.api_key2
                    FROM pt_solar
                        INNER JOIN pt_solar_measuring_info
                        ON pt_solar.id = pt_solar_measuring_info.pt_solar_id
                    WHERE pt_solar.object_no = {0}
                    """
                cursor.execute(sql.format("'%s'" %targetObjectNo))
                result = cursor.fetchall()

                # fetchall 結果がIs not nullの場合
                if result :
                    # タプル1つにSelect句の値が付いている形式で結果値が来るので、Listに変換
                    apiKey = list(result[0])

                    # 計測機器タイプ(10:Eco megane, 20:Laplace, 30:Energy solutions)
                    measuringEquip = apiKey[0]
                    apiKey1 = apiKey[1]
                    apiKey2 = apiKey[2]

                    # add to list
                    responseDataList = []

                    if measuringEquip == LAPLACE_CODE :
                        print("\nAPI Key取得中…")
                        print("\n　◆ ApiKey1 : " + apiKey1)
                        print("　◆ ApiKey2 : " + apiKey2)

                    print("\nAPI処理中…")

                    # create file name
                    cnt = 1
                    if inputMeasuringEquip == LAPLACE_CODE and dataType == SMARTMETER_TYPE :
                        getFileName = f'【LAPLACE Smartmeter】{targetObjectNo}_{targetStartDate}~{targetEndDate}_{cnt} .csv'
                    elif inputMeasuringEquip == LAPLACE_CODE and dataType == PCS_TYPE :
                        getFileName = f'【LAPLACE PCS】{targetObjectNo}_{targetStartDate}~{targetEndDate}_{cnt} .csv'
                    else :
                        getFileName = f'【ENERGYSOLUTIONS】{targetObjectNo}_{targetStartDate}~{targetEndDate}_{cnt} .csv'

                    if inputMeasuringEquip == LAPLACE_CODE and measuringEquip == LAPLACE_CODE :
                        # Api request:smartmeter
                        if dataType == SMARTMETER_TYPE and unit != MONTH :

                            # 1ヶ月以上取得した場合、startDateとendDateのdate_diff.days + 1を31で割った商の値を繰り返し処理回数に設定する
                            if date_diff.days +1 > MONTHLY_DAYS :
                                executeTime = (date_diff.days +1) // MONTHLY_DAYS

                                # 分け前に残りがある場合、繰り返し処理回数1回追加
                                if (date_diff.days +1 / MONTHLY_DAYS) - (date_diff.days // MONTHLY_DAYS) > 0 :
                                    executeTime = executeTime +1

                            # 1か月未満の場合、無条件に1回のみ処理
                            else :
                                executeTime = 1

                            # 1ヶ月未満取得
                            if executeTime == 1 :
                                response = requests.get(
                                    'https://services.energymntr.com/megasolar/' + apiKey1 + '/services/api/download/span.php?&unit=' + unit + '&groupid=1&from=' + targetStartDate + '&to=' + targetEndDate + '&data=measuringdata&format=csv&type=' + dataType,
                                    auth=HTTPDigestAuth(apiKey1, apiKey2),
                                )

                                print("")
                                print(response)

                                if response.status_code == 200 :
                                    contentType = response.headers['Content-Type']
                                    contentDisposition = response.headers['Content-Disposition']
                                    ATTRIBUTE = 'filename='
                                    fileName = contentDisposition[contentDisposition.find(ATTRIBUTE) + len(ATTRIBUTE):]

                                    while os.path.exists(os.path.join(DOWNLOAD_PATH, getFileName)):
                                        cnt += 1
                                        getFileName = f'【LAPLACE Smartmeter】{targetObjectNo}_{targetStartDate}~{targetEndDate}_{cnt} .csv'
                                    # smartmeter dataはそのまま保存 fileName変数未使用
                                    saveFilePath = os.path.join(DOWNLOAD_PATH, getFileName)

                                    with open(saveFilePath, 'wb') as saveFile :
                                        saveFile.write(response.content)

                            # 1ヶ月以上取得
                            else :
                                for count in range(executeTime) :
                                    if count == 0 :
                                        # 最初の終了日は開始日を基点に30日を加算後、EndDateパラメータとして使用
                                        targetEndDateAddDay = targetStartDateToDate + timedelta(days = MONTHLY_DAYS -1)
                                        targetEndDateAddDayFirstToString = targetEndDateAddDay.strftime("%Y%m%d")

                                        response = requests.get(
                                            'https://services.energymntr.com/megasolar/' + apiKey1 + '/services/api/download/span.php?&unit=' + unit + '&groupid=1&from=' + targetStartDate + '&to=' + targetEndDateAddDayFirstToString + '&data=measuringdata&format=csv&type=' + dataType,
                                            auth=HTTPDigestAuth(apiKey1, apiKey2),
                                        )
                                    else :
                                        # 1回以降は終了日+1日を開始日に設定
                                        targetStartDateToDate = targetEndDateAddDay + timedelta(days = 1)
                                        targetStartDateAddDayToString = targetStartDateToDate.strftime("%Y%m%d")

                                        # 上記開始日+30日を終了日に設定
                                        targetEndDateAddDay = targetStartDateToDate + timedelta(days = MONTHLY_DAYS -1)
                                        targetEndDateAddDayToString = targetEndDateAddDay.strftime("%Y%m%d")

                                        if count < executeTime -1 :
                                            response = requests.get(
                                                'https://services.energymntr.com/megasolar/' + apiKey1 + '/services/api/download/span.php?&unit=' + unit + '&groupid=1&from=' + targetStartDateAddDayToString + '&to=' + targetEndDateAddDayToString + '&data=measuringdata&format=csv&type=' + dataType,
                                                auth=HTTPDigestAuth(apiKey1, apiKey2),
                                            )
                                        else :
                                            # 繰り返し処理最後は終了日入力値を endDate として使用するため、条件分岐
                                            response = requests.get(
                                                'https://services.energymntr.com/megasolar/' + apiKey1 + '/services/api/download/span.php?&unit=' + unit + '&groupid=1&from=' + targetStartDateAddDayToString + '&to=' + targetEndDate + '&data=measuringdata&format=csv&type=' + dataType,
                                                auth=HTTPDigestAuth(apiKey1, apiKey2),
                                            )

                                    print("")
                                    print(response)

                                    if response.status_code == 200 :
                                        contentType = response.headers['Content-Type']
                                        contentDisposition = response.headers['Content-Disposition']
                                        ATTRIBUTE = 'filename='
                                        fileName = contentDisposition[contentDisposition.find(ATTRIBUTE) + len(ATTRIBUTE):]

                                        # mergeのための一時ディレクトリの保存
                                        saveFilePath = os.path.join(IMP_PATH, fileName)

                                        with open(saveFilePath, 'wb') as saveFile :
                                            saveFile.write(response.content)

                                # CSVファイル merge(n -> 1) glob関数で targetCsv を collection & read, 空のリストに読み込んだ内容を追加
                                for file in glob.glob(os.path.join(IMP_PATH,'*' + ".csv")) :

                                    # file sizeが0はskip
                                    try :
                                        df = pandas.read_csv(file, encoding = "SHIFT-JIS")
                                    except Exception :
                                        continue

                                    responseDataList.append(df)

                                # concat関数を使用してリストの内容をマージする
                                # axis=0은 수직으로 병합함. axis=1은 수평. ignore_index=True는 Index 値が既存の順序を無視し、順序で並べ替え
                                dataCombine = pandas.concat(responseDataList, axis = 0, ignore_index = False)
                                while os.path.exists(os.path.join(DOWNLOAD_PATH, getFileName)):
                                    cnt += 1
                                    getFileName = f'【LAPLACE Smartmeter】{targetObjectNo}_{targetStartDate}~{targetEndDate}_{cnt} .csv'
                                # to_csv関数として保存します。 Indexを減算するにはFalseに設定
                                dataCombine.to_csv(DOWNLOAD_PATH + getFileName, index = False, encoding = "SHIFT-JIS")

                                # 1day csvファイルとimp directoryの削除
                                if (glob.glob(os.path.join(IMP_PATH, '*' + ".csv"))) :
                                    for file in glob.glob(os.path.join(IMP_PATH, '*' + ".csv")) :
                                        os.remove(file)

                        # Api request:smartmeter and monthly type
                        elif dataType == SMARTMETER_TYPE and unit == MONTH :

                            response = requests.get(
                                'https://services.energymntr.com/megasolar/' + apiKey1 + '/services/api/download/span.php?&unit=' + unit + '&groupid=1&from=' + targetStartDate + '&to=' + targetEndDate + '&data=measuringdata&format=csv&type=' + dataType,
                                auth=HTTPDigestAuth(apiKey1, apiKey2),
                            )

                            print("")
                            print(response)

                            if response.status_code == 200 :
                                contentType = response.headers['Content-Type']
                                contentDisposition = response.headers['Content-Disposition']
                                ATTRIBUTE = 'filename='
                                fileName = contentDisposition[contentDisposition.find(ATTRIBUTE) + len(ATTRIBUTE):]
                                while os.path.exists(os.path.join(DOWNLOAD_PATH, getFileName)):
                                    cnt += 1
                                    getFileName = f'【LAPLACE Smartmeter】{targetObjectNo}_{targetStartDate}~{targetEndDate}_{cnt} .csv'
                                # smartmeter dataはそのまま保存 fileName変数未使用
                                saveFilePath = os.path.join(DOWNLOAD_PATH, getFileName)

                                with open(saveFilePath, 'wb') as saveFile :
                                    saveFile.write(response.content)

                        # Api request:pcs
                        else :
                            for count in range(date_diff.days +1) :

                                # date 文字列変換。 pcsは最大取得期間が1日なので、EndDateまで繰り返し処理
                                targetStartDateAddDay = targetStartDateToDate + timedelta(days = count)
                                targetStartDateAddDayToString = targetStartDateAddDay.strftime("%Y%m%d")

                                response = requests.get(
                                    'https://services.energymntr.com/megasolar/' + apiKey1 + '/services/api/download/daily.php?&unit=' + unit + '&groupid=1&time=' + targetStartDateAddDayToString + '&data=measuringdata&format=csv&type=' + dataType,
                                    auth=HTTPDigestAuth(apiKey1, apiKey2),
                                )

                                print("")
                                print(response)

                                if response.status_code == 200 :
                                    contentType = response.headers['Content-Type']
                                    contentDisposition = response.headers['Content-Disposition']
                                    ATTRIBUTE = 'filename='
                                    fileName = contentDisposition[contentDisposition.find(ATTRIBUTE) + len(ATTRIBUTE):]

                                    # merge のための一時ディレクトリの保存
                                    saveFilePath = os.path.join(IMP_PATH, fileName)

                                    with open(saveFilePath, 'wb') as saveFile :
                                        saveFile.write(response.content)

                            # CSVファイル merge(n -> 1) glob関数で targetCsv を collection & read, 空のリストに読み込んだ内容を追加
                            for file in glob.glob(os.path.join(IMP_PATH,'*' + ".csv")) :

                                # file sizeが0はskip
                                try :
                                    df = pandas.read_csv(file, encoding = "SHIFT-JIS")
                                except Exception :
                                    continue

                                responseDataList.append(df)

                            # concat함수를 이용해서 리스트의 내용을 병합
                            # axis=0은 수직으로 병합함. axis=1은 수평. ignore_index=True는 Index 값이 기존 순서를 무시하고 순서대로 정렬
                            dataCombine = pandas.concat(responseDataList, axis = 0, ignore_index = False)
                            while os.path.exists(os.path.join(DOWNLOAD_PATH, getFileName)):
                                cnt += 1
                                getFileName = f'【LAPLACE PCS】{targetObjectNo}_{targetStartDate}~{targetEndDate}_{cnt} .csv'
                            # to_csv関数として保存します。 Indexを減算するにはFalseに設定
                            dataCombine.to_csv(DOWNLOAD_PATH + getFileName, index = False, encoding = "SHIFT-JIS")

                            # 1day csvファイルとimp directoryの削除
                            if (glob.glob(os.path.join(IMP_PATH, '*' + ".csv"))) :
                                for file in glob.glob(os.path.join(IMP_PATH, '*' + ".csv")) :
                                    os.remove(file)

                        if response.status_code == 401 :
                            print("\nApiKey認証エラー。EPシステムに登録されたApiKeyが正しいか確認する必要があります。")
                        elif response.status_code == 500 :
                            print("\nラプラスサーバーエラーによりAPI処理失敗。")

                    elif inputMeasuringEquip == ENERGYSOLUTIONS_CODE and measuringEquip == ENERGYSOLUTIONS_CODE :

                        for count in range(date_diff.days +1) :

                            # ddos攻撃と同じだから1秒間隔でAPIリクエストを送信...
                            time.sleep(1)

                            # date 文字列変換。 1 Requestあたりの最大取得期間が1日であるため、EndDateまでの繰り返し処理
                            targetStartDateAddDay = targetStartDateToDate + timedelta(days = count)
                            targetStartDateAddDayToString = targetStartDateAddDay.strftime("%Y%m%d")

                            response = requests.get(
                                'https://offgridma-solarmonitor.energy-itsol.com/ejks/pg/g/grPPAApi.aspx?customer_code=' + targetObjectNo + '&target_date=' + targetStartDateAddDayToString + '&teikeisaki_code=19b3d13a7e9640126&customer_code_kubun=2',
                            )

                            print("")
                            print(response)

                            # replace resoponse(byte) to json.
                            if response.status_code == 200 :
                                responseDataList.append(json.loads(response.text))
                            else :
                                print(targetStartDateAddDayToString + "のAPI処理失敗")

                        # 結果がIs not nullの場合
                        if responseDataList :
                            while os.path.exists(os.path.join(DOWNLOAD_PATH, getFileName)):
                                cnt += 1
                                getFileName = f'【ENERGYSOLUTIONS】{targetObjectNo}_{targetStartDate}~{targetEndDate}_{cnt} .csv'
                            # export pandas dataframe to csv
                            pandas.DataFrame(responseDataList).to_csv(path_or_buf = DOWNLOAD_PATH + getFileName, index = False, encoding = "SHIFT-JIS")

                    if (glob.glob(os.path.join(DOWNLOAD_PATH, getFileName + '*'))) :
                        print("\nファイル作成完了")
                        print(" ※ 取得したデータは「C:\python\download\」に保存されています。")

                    elif measuringEquip == ECOMEGANE_CODE :
                        print("\nエコめがねは取得できません。")
                    else :
                        print("\n物件番号に対する計測機器の入力誤りがないか確認してください。")
                else :
                    print("\n入力した物件番号に誤りがないか再度確かめてください。")
        finally :
            # close Database Connection
            db.close()
    os.removedirs(IMP_PATH)

    print("\n続けて取得しますか？")
    print("\n ・0 : プログラム終了")
    print(" ・1 : ラプラスデータを続けて取得")
    print(" ・2 : エナジーソリューションズデータ取得")

    menu_num = input('\n入力: ')

    if menu_num == '1' :
        download()
    elif menu_num == '2':
        if not os.path.exists(DOWNLOAD_PATH) :
            os.makedirs(DOWNLOAD_PATH)
        if not os.path.exists(GET_MULTI_TARGET_PATH) :
            os.makedirs(GET_MULTI_TARGET_PATH)
        if not os.path.exists(ENERGYSOLUTIONS_DOWNDOAD_FILE) :
            os.makedirs(ENERGYSOLUTIONS_DOWNDOAD_FILE)
        chrome_driver()
        os.removedirs(ENERGYSOLUTIONS_DOWNDOAD_FILE)
    else :
        sys.exit()

# エナジーソリューションズのデータ取得
def property_number_list():
    print("\n【エナジーソリューションズデータ取得】")
    print("0. プログラムを終了する")
    print("1. (1件のみ)直接物件番号を入力する")
    print("2. (複数件)【Template】物件番号リストから取得する")

    menu_num = input('\n入力: ')
    if menu_num == '1':
        # レコード番号を入力してデータを抽出
        property_number = input("物件番号を入力してください：")
        print('データ取得期間を入力してください: (例)20200101')
        start_day = input('開始期間 :')
        end_day = input('終了期間 :')

        # データを辞書に格納
        data = {
            "物件番号": [property_number],
            "開始期間": [start_day],
            "終了期間": [end_day]
        }

        # 辞書をDataFrameに変換
        df_data = pandas.DataFrame(data)
        return df_data
    elif menu_num == '2':
        print("\n【複数案件取得手順１】自身の「C:\\python\\target」にアクセスする。（プログラムで既に作成されている）")
        print("【複数案件取得手順２】取得する対象の物件番号を「【Template】物件番号リスト.xlsx」にまとめ、「C:\\python\\target」に入れる。")
        print("【複数案件取得手順３】物件番号リストファイルに暗号化がかかっていないことを確認。")
        input("問題なければEnterを押してください。")

        # Excelファイルのパス
        excel_file_path = GET_MULTI_TARGET_PATH + "【Template】物件番号リスト.xlsx"
        # Excelファイルをデータフレームとして読み込む
        df_data = pandas.read_excel(excel_file_path)
        df_data = df_data.rename(columns={"開始期間 ※入力例) : 20220101": "開始期間", "終了期間 ※入力例) : 20220101": "終了期間"})
        return df_data 
    else:
        print("プログラム終了")
        sys.exit()


def chrome_driver():
    df_data = property_number_list()
    id = os.getenv('ID')
    password = os.getenv('PASS')
    # ダウンロードしたいファイルのURL
    url = "https://offgridma-solarmonitor.energy-itsol.com/ejks/pg/sk/SpLoginPage.aspx"
    # ダウンロード先のフォルダのパス
    download_folder = ENERGYSOLUTIONS_DOWNDOAD_FILE
    # Chromeのオプションを設定
    chrome_options = Options()
    chrome_options.add_experimental_option("prefs", {
        "download.default_directory": download_folder,
        "download.prompt_for_download": False,  # ダウンロード前に確認ダイアログを表示しない
        "download.directory_upgrade": True,
        "safebrowsing.enabled": False
    })
    chrome_options.add_argument('--disable-logging')
    chrome_options.add_experimental_option("excludeSwitches", ["enable-logging"])
    chrome_options.add_argument('--log-level=3')

    # Chromeブラウザを起動
    driver = webdriver.Chrome(options=chrome_options)  # Chromeを準備
    driver.maximize_window()
    # ファイルのURLにアクセス
    driver.get(url)
    time.sleep(5)
    driver.find_element('xpath', "//*[@id='txtUserName']").send_keys(id)  # ID
    time.sleep(2)
    driver.find_element('xpath', "//*[@id='txtPassword']").send_keys(password)  # PASS
    time.sleep(2)
    driver.find_element('xpath', "//*[@id='btnLogin']").click()  # 発電所データダウンロード
    time.sleep(2)

    driver.find_element('xpath', "//*[@id='ctl00_cphMain_btnHatsudenshoDataDL']").click()  # 発電所データダウンロード
    time.sleep(2)
    for i, row in df_data.iterrows():
        property_number = row['物件番号']
        start_day = row['開始期間']
        end_day = row['終了期間']
        merge_start_date = datetime.strptime(str(start_day), '%Y%m%d').strftime('%Y%m%d')
        merge_end_date = datetime.strptime(str(end_day), '%Y%m%d').strftime('%Y%m%d')
        start_date = datetime.strptime(str(start_day), '%Y%m%d')
        end_date = datetime.strptime(str(end_day), '%Y%m%d')

        # 60日ごとに期間を分割してデータ取得
        while start_date <= end_date:
            next_end_date = min(start_date + timedelta(days=59), end_date)
            download_data(driver, start_date.strftime('%Y%m%d'), next_end_date.strftime('%Y%m%d'), property_number, download_folder)
            start_date = next_end_date + timedelta(days=1)
        csv_files = glob.glob(os.path.join(download_folder, '*.csv'))
        # CSVファイルの中身を追加していくリストを用意
        data_list = []

        # 読み込むファイルのリストを走査し、データをリストに追加
        for file in csv_files:
            data_list.append(pandas.read_csv(file, encoding='shift-jis'))

        # リストを全て行方向に結合
        merge_df = pandas.concat(data_list, axis=0)
        # 最初の行を取得
        first_row = merge_df.iloc[0]

        # 各列の値を変数に代入
        hatsuden_value = first_row['発電所']
        # ファイルの削除
        for file in csv_files:
            os.remove(file)

        cnt = 1
        # 目的のファイル名
        merge_file_name = f"【エナソリュ】{property_number}_{hatsuden_value}_{merge_start_date.replace('/', '')}_{merge_end_date.replace('/', '')}_{cnt}.xlsx"
        # 移動先のフォルダのパス
        destination_folder = DOWNLOAD_PATH
        while os.path.exists(os.path.join(destination_folder, merge_file_name)):
            cnt += 1
            merge_file_name = f"【エナソリュ】{property_number}_{hatsuden_value}_{merge_start_date.replace('/', '')}_{merge_end_date.replace('/', '')}_{cnt}.xlsx"
        merge_file_path = os.path.join(destination_folder, merge_file_name)

        merge_df.to_excel(merge_file_path, index=False)        
        excel_create(merge_file_path)

    print('作成ファイルは「C:\\python\\download\\」にあります')
    driver.quit()  # ブラウザを閉じる

def download_data(driver, start_date, end_date, property_number, download_folder):
    start_day_element = driver.find_element(By.ID, "ctl00_cphMain_txtCSVDateFrom")  # 開始
    start_day_element.clear()  # 開始
    start_day_element.send_keys(start_date)  # 開始
    time.sleep(2)
    end_day_element = driver.find_element(By.ID, "ctl00_cphMain_txtCSVDateTo")  # 終了
    end_day_element.clear()  # 終了
    end_day_element.send_keys(end_date)
    time.sleep(2)
    property_number_element = driver.find_element(By.ID, "ctl00_cphMain_txtHatsudenshoBango") # 物件番号入力
    property_number_element.clear() # 物件番号入力
    property_number_element.send_keys(property_number) # 物件番号入力
    time.sleep(2)    
    select_data_format = driver.find_element(By.ID, "ctl00_cphMain_ddlDataKeishiki")  # データ形式
    select = Select(select_data_format)
    select.select_by_visible_text('累計値・差分')
    time.sleep(2)
    select_data_unit = driver.find_element(By.ID, "ctl00_cphMain_ddlDataShukeiTani")  # データ集計単位
    select2 = Select(select_data_unit)
    select2.select_by_visible_text('30分単位')
    time.sleep(2)

    driver.find_element('xpath', '//*[@id="ctl00_cphMain_btnRawShohichiDL"]').click()  # データ出力
    time.sleep(5)
    # ダウンロード後のファイルパス
    downloaded_file_path = os.path.join(download_folder, "hatsudenshoshohidata.csv")

    # 目的のファイル名
    file_name = f"{start_date.replace('/', '')}_{end_date.replace('/', '')}_{property_number}_エナジーソリューションズファイル確認.csv"
    new_file_path = os.path.join(download_folder, file_name)

    # ファイル名を変更する
    os.rename(downloaded_file_path, new_file_path)

def excel_create(merge_file_path):
    #デザイン修正
    wb = load_workbook(merge_file_path)
    ws = wb.active

    def column_width():
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 7
        ws.column_dimensions['E'].width = 25
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 20
        ws.column_dimensions['I'].width = 20
        ws.column_dimensions['J'].width = 20
        ws.column_dimensions['K'].width = 20
        ws.column_dimensions['L'].width = 25
        ws.column_dimensions['M'].width = 25
        ws.column_dimensions['N'].width = 25
        ws.column_dimensions['O'].width = 25
        ws.column_dimensions['P'].width = 15
        ws.column_dimensions['Q'].width = 15
        ws.column_dimensions['R'].width = 15
        ws.column_dimensions['S'].width = 15
        ws.column_dimensions['T'].width = 20
        ws.column_dimensions['U'].width = 25

    def cell_color():
        range=ws['A1':'U1']
        for row in range:
            for cell in row:
                cell.fill = PatternFill(fill_type='solid', fgColor=Color('FCD5B4'))

    def line_borders(ws):
        box = Border(left=Side(border_style='thin', color='FF000000'),
                    right=Side(border_style='thin', color='FF000000'),
                    top=Side(border_style='thin', color='FF000000'),
                    bottom=Side(border_style='thin', color='FF000000'))

        for  num, row in enumerate(ws.rows):
            for cell in row:
                cell.border=box

    def font_set():
        font_m = Font(name='Meiryo UI', size=11)
        for  num, row in enumerate(ws.rows):
            for cell in row:
                cell.font=font_m

    def underline():
        range=ws['A1':'U1']
        for row in range:
            for cell in row:
                cell.font = Font(bold=True)
                cell.font = Font(u='double')

    column_width()
    cell_color()
    line_borders(ws)
    underline()
    font_set()
    # line_borders(ws)
    wb.save(merge_file_path)
    print("ファイル生成完了\n")

def main():
    print("\nどちらのデータを取得するか選択してください")
    print("\n1.ラプラス")
    print("\n2.エナジーソリューションズ")
    print("\n0.終了する場合")
    print("\n※1の場合でもAPIを使ってエナジーソリューションズのデータ取得は可能です")
    first_choice = int(input("\n ・1 or 2を入力 : "))
    try:
        if first_choice == 1:
            # create local directory
            if not os.path.exists(DOWNLOAD_PATH) :
                os.makedirs(DOWNLOAD_PATH)
            if not os.path.exists(IMP_PATH) :
                os.makedirs(IMP_PATH)
            if not os.path.exists(GET_MULTI_TARGET_PATH) :
                os.makedirs(GET_MULTI_TARGET_PATH)
            download()
        elif first_choice == 2:
            # create local directory
            if not os.path.exists(DOWNLOAD_PATH) :
                os.makedirs(DOWNLOAD_PATH)
            if not os.path.exists(GET_MULTI_TARGET_PATH) :
                os.makedirs(GET_MULTI_TARGET_PATH)
            if not os.path.exists(ENERGYSOLUTIONS_DOWNDOAD_FILE) :
                os.makedirs(ENERGYSOLUTIONS_DOWNDOAD_FILE)
            chrome_driver()
            os.removedirs(ENERGYSOLUTIONS_DOWNDOAD_FILE)
        else:
            print("プログラム終了")
            sys.exit()
        main()
    except Exception as e:
        print(f'エラー: {e}')
        logger.error(f'エラー: {e}', exc_info=True)
        input('ログの共有お願いいたします、一旦Enterを押して閉じます')
        sys.exit()

if __name__ == '__main__':
     main()


